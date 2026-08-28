import unittest
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from peripatetic_timetable.exports import available_export_copy, export_excel
from peripatetic_timetable.repository import TimetableRepository


class ExcelExportTests(unittest.TestCase):
    def test_locked_export_uses_a_unique_dated_copy_name(self):
        with TemporaryDirectory() as directory:
            selected = Path(directory) / "peripatetic_timetable.xlsx"
            expected = Path(directory) / "peripatetic_timetable_2026-08-28_133700.xlsx"
            expected.touch()
            self.assertEqual(
                available_export_copy(selected, datetime(2026, 8, 28, 13, 37)),
                Path(directory) / "peripatetic_timetable_2026-08-28_133700_2.xlsx",
            )

    def test_excel_export_contains_all_reports_and_complete_timetable(self):
        timetable = TimetableRepository().load_baseline()
        with TemporaryDirectory() as directory:
            path = Path(directory) / "timetable.xlsx"
            export_excel(timetable, path)
            workbook = load_workbook(path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "Timetable",
                        "Teacher Movement",
                        "Coverage Audit",
                        "Policy Audit",
                        "Staffing Notes",
                        "Change Log",
                    ],
                )
                sheet = workbook["Timetable"]
                self.assertEqual(sheet["A4"].value, "Attard")
                self.assertEqual(sheet["B4"].value, "20 classes")
                self.assertIn("Rebecca Bonello", sheet["C4"].value)
                self.assertEqual(sheet.max_row, 9)
                self.assertEqual(sheet.max_column, 7)
                self.assertEqual(sheet.page_setup.orientation, "landscape")
                self.assertEqual(sheet.page_setup.paperSize, 9)
                self.assertEqual(sheet.page_setup.fitToWidth, 1)
                self.assertEqual(sheet.page_setup.fitToHeight, 1)
                self.assertTrue(sheet.sheet_properties.pageSetUpPr.fitToPage)
                self.assertEqual(str(sheet.print_area), "'Timetable'!$A$1:$G$9")
                self.assertLessEqual(sheet.page_margins.left, 0.2)
                self.assertLessEqual(sheet.page_margins.right, 0.2)
                self.assertTrue(sheet.print_options.horizontalCentered)
                self.assertTrue(sheet.print_options.verticalCentered)

                report_sheet = workbook["Teacher Movement"]
                self.assertEqual(report_sheet.page_setup.orientation, "landscape")
                self.assertEqual(report_sheet.page_setup.paperSize, 9)
                self.assertEqual(report_sheet.page_setup.fitToWidth, 1)
                self.assertEqual(report_sheet.page_setup.fitToHeight, 0)
                self.assertEqual(
                    str(report_sheet.print_title_rows),
                    "$1:$1",
                )
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
