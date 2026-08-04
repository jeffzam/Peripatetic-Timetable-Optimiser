import csv
from pathlib import Path

from .audit import coverage_audit
from .config import DAYS

def export_csv(data: dict, path: str | Path) -> None:
    with Path(path).open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["school", "day", "subject", "teacher", "baseline"])
        writer.writeheader()
        writer.writerows(data["assignments"])

def export_excel(data: dict, path: str | Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timetable"
    dark, light = PatternFill("solid", fgColor="17324D"), PatternFill("solid", fgColor="E8F0F8")
    white = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="3B4A54")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Peripatetic Teachers Timetable"
    sheet["A1"].font, sheet["A1"].fill = Font(size=16, bold=True, color="FFFFFF"), dark
    sheet["A1"].alignment = Alignment(horizontal="center")
    for column, header in enumerate(["PRIMARY SCHOOLS", "NUMBER OF CLASSES", *DAYS], 1):
        cell = sheet.cell(3, column, header)
        cell.fill, cell.font, cell.border = dark, white, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row, school in enumerate(data["schools"], 4):
        values = [f"{school['name']} ({school['classes']})", school["breakdown"]]
        values += ["\n".join(f"{item['subject']} - {item['teacher']}" for item in data["assignments"]
                             if item["school"] == school["name"] and item["day"] == day)
                   for day in DAYS]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.border, cell.alignment = border, Alignment(vertical="top", wrap_text=True)
            if column == 1:
                cell.font, cell.fill = Font(bold=True), light
        sheet.row_dimensions[row].height = 95
    for column, width in enumerate([22, 20, 32, 32, 32, 32, 32], 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    audit = workbook.create_sheet("Teacher Coverage Audit")
    audit.append(["Teacher", "Days Count", "Assigned Days", "Missing Days", "Schools", "Status"])
    for row in coverage_audit(data):
        audit.append([row["teacher"], row["days_count"], ", ".join(row["assigned_days"]),
                      ", ".join(row["missing_days"]), ", ".join(row["schools"]), row["status"]])
    workbook.save(path)
