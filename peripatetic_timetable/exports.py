import csv
from pathlib import Path

from .audit import audit_timetable
from .config import DAYS
from .domain import Timetable
from .reports import coverage_report, movement_matrix

def export_csv(timetable: Timetable, path: str | Path) -> None:
    with Path(path).open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["School", "Day", "Subject", "Teacher", "Baseline"])
        for item in timetable.assignments:
            writer.writerow([item.school, item.day, item.subject, item.teacher, item.baseline])

def export_excel(timetable: Timetable, path: str | Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    def configure_a4_landscape(
        worksheet,
        *,
        fit_to_height: int,
        print_area: str | None = None,
        repeat_header_rows: str | None = None,
        center_vertically: bool = False,
    ) -> None:
        """Apply predictable, printer-friendly settings to an exported sheet."""
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_setup.scale = None
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = fit_to_height
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
        worksheet.page_margins = PageMargins(
            left=0.2,
            right=0.2,
            top=0.35,
            bottom=0.35,
            header=0.1,
            footer=0.1,
        )
        worksheet.print_options.horizontalCentered = True
        worksheet.print_options.verticalCentered = center_vertically
        worksheet.sheet_view.showGridLines = False
        if print_area is not None:
            worksheet.print_area = print_area
        if repeat_header_rows is not None:
            worksheet.print_title_rows = repeat_header_rows

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timetable"
    navy = PatternFill("solid", fgColor="17324D")
    pale_blue = PatternFill("solid", fgColor="E8F0F8")
    white_bold = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="8EA0B2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Peripatetic Teachers' Timetable"
    sheet["A1"].font = Font(size=17, bold=True, color="FFFFFF")
    sheet["A1"].fill = navy
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    headers = ["Primary school", "Classes", *DAYS]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(3, column, header)
        cell.fill, cell.font, cell.border = navy, white_bold, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number, school in enumerate(timetable.schools, 4):
        values = [school.name, f"{school.classes} classes"]
        for day in DAYS:
            values.append("\n".join(
                f"{item.subject} — {item.teacher}"
                for item in timetable.assignments_for(school=school.name, day=day)
            ))
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == 1:
                cell.fill, cell.font = pale_blue, Font(bold=True)
        busiest_day = max(
            len(timetable.assignments_for(school=school.name, day=day))
            for day in DAYS
        )
        sheet.row_dimensions[row_number].height = max(45, busiest_day * 18)
    for column, width in enumerate([20, 16, 32, 32, 32, 32, 32], 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "C4"
    last_timetable_row = 3 + len(timetable.schools)
    configure_a4_landscape(
        sheet,
        fit_to_height=1,
        print_area=f"A1:G{last_timetable_row}",
        center_vertically=True,
    )

    movement = workbook.create_sheet("Teacher Movement")
    movement.append(["Teacher", *DAYS])
    for teacher, days in movement_matrix(timetable).items():
        movement.append([teacher, *(days[day] for day in DAYS)])

    coverage = workbook.create_sheet("Coverage Audit")
    coverage.append(["Teacher", "Days", "Assigned days", "Missing days", "Conflict days", "Schools", "Status"])
    for row in coverage_report(timetable):
        coverage.append([
            row.teacher, len(row.assigned_days), ", ".join(row.assigned_days),
            ", ".join(row.missing_days), ", ".join(row.conflict_days),
            ", ".join(row.schools), row.status,
        ])

    audit = workbook.create_sheet("Policy Audit")
    audit.append(["Level", "Check", "Teacher", "School", "Day", "Detail"])
    for issue in audit_timetable(timetable):
        audit.append([
            issue.severity.value,
            issue.code,
            issue.teacher,
            issue.school,
            issue.day,
            issue.detail,
        ])

    staffing = workbook.create_sheet("Staffing Notes")
    staffing.append(["Person or post", "Status", "Note"])
    for note in timetable.staff_notes:
        staffing.append([note.name, note.status, note.note])

    changes = workbook.create_sheet("Change Log")
    changes.append(["Version", "Note"])
    for entry in timetable.change_log:
        changes.append([entry.version, entry.note])

    for worksheet in workbook.worksheets[1:]:
        for cell in worksheet[1]:
            cell.fill, cell.font, cell.border = navy, white_bold, border
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            letter = get_column_letter(column_cells[0].column)
            maximum = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[letter].width = min(max(maximum + 2, 12), 55)
        configure_a4_landscape(
            worksheet,
            fit_to_height=0,
            repeat_header_rows="1:1",
        )
    workbook.save(path)
